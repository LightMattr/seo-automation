import sys

import googlesearch
import requests
import bs4
import webbrowser

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter


query = input('Enter a search term: ')


def main(): # runs main program
    top_results = get_link()
    results_list = get_title(top_results)
    create_sheet(results_list)


def get_link():  # gets links for first page results, and appends it to results list
    search = googlesearch.search(query, tld='com', num=10, stop=1, pause=2)
    results = []
    for link in search:
        results.append(link)
    return results


def get_title(results): # Gets title from top search results & stores in list of dicts
    result_dicts = []
    position = 1
    for link in results:
        page_info = {}
        site = requests.get(link)
        format = bs4.BeautifulSoup(site.text, 'lxml')
        page_title = format.select('title')
        try:
            page_info['title'] = page_title[0].getText()
        except Exception as e:
            page_info['title'] = '<<NO TITLE FOUND>>'
        page_info['keyword'] = query
        page_info['position'] = position
        page_info['link'] = link
        position += 1
        result_dicts.append(page_info)
    return result_dicts


def create_sheet(page_dicts): # Creates a workbook called Search_Results and prints search result data on each row
    wb = Workbook()
    dest_file = 'Search_Results.xlsx'
    ws1 = wb.active
    rowx = 1
    max_row = len(page_dicts)
    for _ in range(max_row):
        ws1.cell(row=rowx, column=1, value=page_dicts[rowx-1]['keyword'])
        ws1.cell(row=rowx, column=2, value=page_dicts[rowx-1]['title'])
        ws1.cell(row=rowx, column=3, value=page_dicts[rowx-1]['position'])
        ws1.cell(row=rowx, column=4, value=page_dicts[rowx-1]['link'])
        rowx += 1
    wb.save(filename = dest_file)
    wb.close()
    print("Your file is ready")


if __name__ == '__main__':
    main()
